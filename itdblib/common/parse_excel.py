# !/usr/bin/env python
# -*- coding: utf-8 -*-
# author: wangzq <wangzhenqing1008@163.com>


from qg.core import log as logging
from openpyxl import load_workbook
from itdblib.common.asset_info_utils import AssetInfoUtils
from itdblib.common.type_info_utils import (
    wireList,
    vpnList,
    macList
)

LOG = logging.getLogger(__name__)


def format_field_map():
    aiu = AssetInfoUtils()
    chlist = aiu.input_excel_ch_list()
    enlist = aiu.input_excel_en_list()
    e_len = len(enlist)
    return dict([(chlist[i], enlist[i]) for i in range(e_len)])


def format_phone_field_map():
    aiu = AssetInfoUtils()
    chlist = aiu.input_phone_excel_ch_list()
    enlist = aiu.input_phone_excel_en_list()
    e_len = len(enlist)
    return dict([(chlist[i], enlist[i]) for i in range(e_len)])


def data_set_from_excel(fileFullPath, sheetName='assets'):
    """ read and set data from excel
    """
    fieldMap = format_field_map()
    try:
        wb = load_workbook(filename=fileFullPath)
        sheet = wb[sheetName]
        head = map(lambda x: x.value, sheet.rows[0])
        dataSet = []
        for row in sheet.rows[1:]:
            data_row = {}
            for i, cell in enumerate(row):
                data_row[fieldMap[head[i]]] = str(cell.value).strip() if cell.value else ''
            if data_row['asset_type'] in macList():
                data_row['type_info'] = dict(mac=data_row['mac'])
            elif data_row['asset_type'] in wireList():
                data_row['type_info'] = dict(imei=data_row['imei'],
                                             version=data_row['version'])
            elif data_row['asset_type'] in vpnList():
                data_row['type_info'] = dict(userful_life=data_row['userful_life'])
            else:
                data_row['type_info'] = {}

            dataSet.append(data_row)
        return True, dataSet
    except Exception, _ex:
        LOG.error("error occured while read data from excel : %s" % str(_ex))
        return False, []


def get_phone_data_set_from_excel(fileFullPath, sheetName='phone'):
    """ read and set data from excel
    """
    fieldMap = format_phone_field_map()
    try:
        wb = load_workbook(filename=fileFullPath)
        sheet = wb[sheetName]
        head = map(lambda x: x.value, sheet.rows[0])
        dataSet = []
        for row in sheet.rows[1:]:
            data_row = {}
            for i, cell in enumerate(row):
                data_row[fieldMap[head[i]]] = str(cell.value).strip() if cell.value else ''
            dataSet.append(data_row)
        return True, dataSet
    except Exception, _ex:
        LOG.error("error occured while read data from excel : %s" % str(_ex))
        return False, []


def get_excel_sheet_name(fileFullPath):
    wb = load_workbook(filename=fileFullPath)
    sheetNames = ','.join(wb.get_sheet_names())
    LOG.info('Available sheets are %s' % sheetNames)
    if 'assets' in sheetNames and 'phone' in sheetNames:
        return True, 3
    if 'assets' in sheetNames:
        return True, 1
    if 'phone' in sheetNames:
        return True, 2
    return False, 'Sheet为'+sheetNames+', 不包含assets或者phone, 请检查!'